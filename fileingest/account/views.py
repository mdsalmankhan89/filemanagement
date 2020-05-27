from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.models import User, auth
from django.core.files.storage import FileSystemStorage
import json
import re
import os
import openpyxl
from django.conf import settings
from django.conf.urls.static import static
from django.core import serializers
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
from .forms import FileForm
from .models import Files, FileLogs, FilesData, Rules
import psycopg2
import xlrd

import boto3
from botocore.exceptions import NoCredentialsError

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated  # <-- Here

from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status

from .serializers import FileSerializer

class HelloView(APIView):
	permission_classes = (IsAuthenticated,)

	def get(self,request):
		content = {'get':'val'}
		return Response(content)

	def post(self,request, *args, **kwargs):
		myvar = request.user.email
		content = request.data
		return Response(content)

class FileView(APIView):
	parser_classes = (MultiPartParser, FormParser)
	def post(self, request, *args, **kwargs):
		#handle_files()
		request.data["user"]=request.user.id

		file_serializer = FileSerializer(data=request.data)
		if file_serializer.is_valid():
			file_serializer.save()

			#print(request.data['module'])

			filename =os.path.basename(file_serializer.data["files"])	
			user = file_serializer.data["user"]
			uploadid = file_serializer.data["uploadid"]
			module = request.data['module']
			handle_filesAPI(filename,user,module,uploadid)
			#print(file_serializer.uploadid)
			return Response(file_serializer.data, status=status.HTTP_201_CREATED)
		else:
			return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Create your views here.
def login(request):
	if request.method == 'POST':
		username = request.POST['username']
		password = request.POST['password']
		
		user = auth.authenticate(username=username, password=password)
		
		if user is not None:
			auth.login(request, user)
			return redirect("upload")
		else:
			messages.info(request,"Invalid credentials")
			return redirect('login')
	else:
		return render(request, 'login.html')
		
def logout(request):
	auth.logout(request)
	return redirect("/")
	
def upload(request):

	#filemetas = FilesData.objects.filter(userid=request.user.id).values('filelabel')
	filemetas = FilesData.objects.filter(userid=request.user.id).values('filelabel').distinct()
	
	try:
		filelist = Files.objects.filter(user_id=request.user)
	except Files.DoesNotExist:
		filelist = None
	
	fileloglist = FileLogs.objects.all()
	
	if request.method == 'POST':
		form = FileForm(request.POST, request.FILES)
		if form.is_valid():
			module = request.POST.get('module')
			handle_files(request.FILES['files'],request.user, module)
			filelist = Files.objects.filter(user_id=request.user)
			
			return render(request, 'upload.html', { 'form' : form, 'filelist' : filelist, 'fileloglist' : fileloglist, 'filemetas' : filemetas  })
	else:
		form = FileForm()
	return render(request, 'upload.html', { 'form' : form, 'filelist' : filelist, 'fileloglist' : fileloglist, 'filemetas' : filemetas })
	
def handle_filesAPI(excel_file, user, module,uploadid):
	entry = FileLogs(uploadid=Files.objects.get(uploadid=uploadid), files=excel_file, logs='File Received')
	entry.save()
	
	validate(excel_file, excel_file, uploadid, module)

def handle_files(excel_file, user, module):

   newfile = Files(files=excel_file,user=user)
   newfile.save()

   entry = FileLogs(uploadid=Files.objects.get(uploadid=newfile.uploadid), files=excel_file, logs='File Received')
   entry.save()
   
   validate(excel_file, newfile.files.url, newfile.uploadid, module)

def validate(excel_file, new_file, uploadid, module):
	
	print(excel_file)
	#filename = "AP&ISA_202002Feb_Maximo_Resoln_wos_BI_20200310.xlsx"
	filename = str(excel_file)
	BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
	FILES_FOLDER = os.path.join(BASE_DIR,'media\\files')
	filepath = os.path.join(FILES_FOLDER,filename)
	#filepath = "C:\\Users\\salman\\Desktop\\DjangoLearning\\projects\\DataIngestion\\filemanagement\\fileingest\\media\\files\\APISA_202002Feb_Maximo_Resoln_wos_BI_20200310.xlsx"
		
	rules = Rules.objects.all()
	
	validationResult  = {}

	validationResult = validateFileExtension(filepath,validationResult)

	if validationResult["extension"] == True:
        
		validationResult = validateFileNamePattern(rules,filepath,validationResult,module)
		entry = FileLogs(uploadid=Files.objects.get(uploadid=uploadid), files=excel_file, logs='File Extention Validation : Successful')
		entry.save()
		print(validationResult)

		if validationResult["filename_pattern"][2]== True:

			validationResult = validatePeriodPattern(rules,filepath,validationResult,module)
			entry = FileLogs(uploadid=Files.objects.get(uploadid=uploadid), files=excel_file, logs='File Name Validation : Successful')
			entry.save()
			print(validationResult)			

			if validationResult["period_pattern"][3] == True:

				validationResult = validateSheets(rules,filepath,validationResult,module)
				entry = FileLogs(uploadid=Files.objects.get(uploadid=uploadid), files=excel_file, logs='Period File Validation : Successful')
				entry.save()
				print(validationResult)				

				if validationResult["sheetname_pattern"][1] ==True:

					validationResult=validateColumns(rules,filepath,validationResult,module)
					entry = FileLogs(uploadid=Files.objects.get(uploadid=uploadid), files=excel_file, logs=('Sheet Name Validation : Successful'))
					entry.save()
					print(validationResult)					

					if validationResult["columns"][1] == True:
						entry = FileLogs(uploadid=Files.objects.get(uploadid=uploadid), files=excel_file, logs=('Columns Validation : Successful'))
						entry.save()
						print ("success")
						exceltocsv(filepath,FILES_FOLDER,module)
						
					else:
						entry = FileLogs(uploadid=Files.objects.get(uploadid=uploadid), files=excel_file, logs='Columns Validation : Failed')
						entry.save()
						
				else:
					entry = FileLogs(uploadid=Files.objects.get(uploadid=uploadid), files=excel_file, logs='Sheet Name Validation : Failed')
					entry.save()
						
			else:
				entry = FileLogs(uploadid=Files.objects.get(uploadid=uploadid), files=excel_file, logs='Period File Validation : Failed')
				entry.save()
						
		else:
			entry = FileLogs(uploadid=Files.objects.get(uploadid=uploadid), files=excel_file, logs='File Name Validation : Failed')
			entry.save()
	
	else:
		entry = FileLogs(uploadid=Files.objects.get(uploadid=uploadid), files=excel_file, logs='File Extention Validation : Failed')
		entry.save()

def exceltocsv(filepath,FILES_FOLDER,module):
	df = pd.read_excel(filepath, sheet_name=None)
	somevar = FILES_FOLDER
	csvlist = []
	for key, value in df.items():
		df[key].to_csv('%s\\%s.csv' %(somevar ,key),index =None, header= True)	 
		csvlist.append('%s\\%s.csv' %(somevar ,key))
	uploadtordbms(csvlist,module)
 
def uploadtordbms(csvlist,module):
	print(csvlist)
	ACCESS_KEY = '' 
	SECRET_KEY = ''
	modules3 = module+'/{}'
	for filepath in csvlist:
		print(filepath)
		local_file = filepath
		bucket_name = ''
		print(bucket_name)
		s3_file_name = os.path.basename(local_file)

		s3 = boto3.client('s3', aws_access_key_id=ACCESS_KEY,aws_secret_access_key=SECRET_KEY)
		
		try:
			s3.upload_file(local_file, bucket_name,modules3.format(s3_file_name))
			print("Upload Successful")
		except FileNotFoundError:
			print("The file was not found")
			return False
		except NoCredentialsError:
			print("Credentials not available")
			return False

def validateFileExtension(filepath,validationResult):

	extension  = os.path.splitext(filepath)[1]

	#check if extension is correct
	if re.search(".xlsx",extension,re.IGNORECASE):
		validationResult.update({'extension':True})
	else:
		validationResult.update({'extension':False})

	return validationResult

def validateFileNamePattern(rules,filepath,validationResult,module):
	filename = os.path.basename(filepath)
	validationResult.update({'filename_pattern':["",filename,False]})
    
	for rule in rules:
		ruleJSON = json.loads(rule.rule)
		if re.search(ruleJSON["filename_pattern"],filename,re.IGNORECASE) and re.search(ruleJSON["module"],module,re.IGNORECASE):
			validationResult.update({'filename_pattern':[ruleJSON["filename_pattern"],filename,True]})
    
	return validationResult

def validatePeriodPattern(rules,filepath,validationResult,module):
	filename = os.path.basename(filepath)
	validationResult.update({'period_pattern':["",filename,"",False]})

	for rule in rules:
		ruleJSON = json.loads(rule.rule)

		if re.search(ruleJSON["filename_pattern"],filename,re.IGNORECASE) and re.search(ruleJSON["module"],module,re.IGNORECASE):

			if re.search(ruleJSON["period_pattern"],filename,re.IGNORECASE):
				validationResult.update({'period_pattern':[ruleJSON["period_pattern"],filename,(re.search(ruleJSON["period_pattern"],filename,re.IGNORECASE)).group(0),True]})
				break
			else:
				validationResult.update({'period_pattern':[ruleJSON["period_pattern"],filename,"",False]})

	return validationResult

def validateSheets(rules,filepath,validationResult,module):
	expectedsheets = {}

	book = openpyxl.load_workbook(filepath)
	presentSheets = [x.lower() for x in book.sheetnames]

	for rule in rules:
		ruleJSON = json.loads(rule.rule)

		if (re.search(ruleJSON["module"],module,re.IGNORECASE) and (ruleJSON["period_pattern"]==validationResult["period_pattern"][0]) and re.search(ruleJSON["filename_pattern"],validationResult["filename_pattern"][0],re.IGNORECASE)):
			if ([item for item in [(ruleJSON["sheetname_pattern"]).lower()] if item in presentSheets]):
				expectedsheets.update({ruleJSON["sheetname_pattern"]:True})
			else:
				expectedsheets.update({ruleJSON["sheetname_pattern"]:False})
	if ([item for item in [False] if item in expectedsheets.values()] or (not expectedsheets)):
		validationResult.update({'sheetname_pattern':[expectedsheets,False]})
	else:
		validationResult.update({'sheetname_pattern':[expectedsheets,True]})
    
	return validationResult

def validateColumns(rules,filepath,validationResult,module):
	expectedresult = {}
	book = openpyxl.load_workbook(filepath) 

	for sheetname in book.sheetnames:
		partialmatch_Header=0 
		partialmatch_Counter =0 
		partial_delta = []
		print(sheetname)
		xl_sheet = book[sheetname]
		print(xl_sheet)
		
		for rule in rules:
			ruleJSON = json.loads(rule.rule)
			ruleid = rule.ruleid
			
			
			if (re.search(ruleJSON["module"],module,re.IGNORECASE) and (ruleJSON["period_pattern"]==validationResult["period_pattern"][0]) and re.search(ruleJSON["filename_pattern"],validationResult["filename_pattern"][0],re.IGNORECASE) and re.search(ruleJSON["sheetname_pattern"],sheetname,re.IGNORECASE)):
				print("Entry")
				if ruleJSON["header"] and ruleJSON["column_offset"]:
					columns=[]
					for cell in xl_sheet[ruleJSON["header"]]:
						columns.append(cell.value)
					missing_columns =  [item for item in ruleJSON["columns"] if item not in columns]

					if missing_columns:
						expectedresult.update({sheetname:[ruleJSON["columns"],missing_columns,ruleJSON["header"],ruleJSON["column_offset"],False,None]})
					else:
						expectedresult.update({sheetname:[ruleJSON["columns"],None,ruleJSON["header"],ruleJSON["column_offset"],True,ruleid]})
						if int(ruleJSON["column_offset"]) > 0:
							xl_sheet.delete_cols(1, int(ruleJSON["column_offset"]))
						if int(ruleJSON["header"])>1 :
							xl_sheet.delete_rows(1,(int(ruleJSON["header"])-1))
                
				if (ruleJSON["header"] == "" or ruleJSON["header"] is None):
					no_of_rows = xl_sheet.max_row
					counter = 1
					while counter < no_of_rows:
						columns=[]
						for cell in xl_sheet[counter]:
							columns.append(cell.value)
							
						if [item for item in [ruleJSON["columns"]] if item in columns]:
							if counter==1:
								partialmatch_Counter = len([item for item in [ruleJSON["columns"]] if item in columns])
								partialmatch_Header = counter
								partial_delta = [item for item in [ruleJSON["columns"]] if item in columns]
								partial_col_offset = 0
								for val in columns:
									if val is None:
										partial_col_offset += 1
							elif  partialmatch_Counter > len([item for item in [ruleJSON["columns"]] if item in columns]):
								partialmatch_Counter = len([item for item in [ruleJSON["columns"]] if item in columns])
								partialmatch_Header = counter
								partial_delta = [item for item in [ruleJSON["columns"]] if item in columns]
								partial_col_offset = 0
								for val in columns:
									if val is None:
										partial_col_offset += 1
                                        
							counter = counter + 1
							if counter > 50:
								break
							continue
						else:
							if ruleJSON["column_offset"]:
								if col_offset > 0:
									xl_sheet.delete_cols(1, ruleJSON["column_offset"])
								if counter > 1 :
									xl_sheet.delete_rows(1,(counter-1))
								expectedresult.update({sheetname:[ruleJSON["columns"],None,counter,ruleJSON["column_offset"],True,ruleid]})
                                
							else:
								col_offset = 0
								for val in columns:
									if val is None:
										col_offset += 1
								if col_offset > 0:
									xl_sheet.delete_cols(1, col_offset)
								if counter > 1 :
									xl_sheet.delete_rows(1,(counter-1))
								expectedresult.update({sheetname:[ruleJSON["columns"],None,counter,col_offset,True,ruleid]})
								break

					if counter > 50:
						print (partialmatch_Counter)
						print (partialmatch_Header)
						print (partial_delta )
						print ("header not found")
						expectedresult.update({sheetname:[ruleJSON["columns"],partial_delta,partialmatch_Header,partial_col_offset,False,None]})

	print(expectedresult.values())
	result =[]
	for resvals in expectedresult.values():
		result.append(resvals[4])
    
	if ([item for item in [False] if item in result] or (not result)):
		validationResult.update({'columns':[expectedresult,False]})
	else:
		validationResult.update({'columns':[expectedresult,True]})

	return validationResult
			
def register(request):
	
	rules = Rules.objects.all()
	ruleList=[]
	for rule in rules:
		RuleLabel = json.loads(rule.rule)
		ruleList.append(RuleLabel["module"])
	
	ruleList = list(set(ruleList))
	
	if request.method == 'POST':
		firstname = request.POST['firstname']
		lastname = request.POST['lastname']
		username = request.POST['username']
		email = request.POST['email']
		password = request.POST['password']
		confirmpassword = request.POST['confirmpassword']
		filelabel = []
		filelabel= request.POST.getlist('mdname')	  
		
		if username=='':
			messages.info(request, 'Please provide Username')
			return redirect('register')
		elif password=='':
			messages.info(request, 'Please provide password')
			return redirect('register')
		elif password==confirmpassword:
			if User.objects.filter(username=username).exists():
				messages.info(request, 'Username Taken')
				return redirect('register')
			elif User.objects.filter(email=email).exists():
				messages.info(request, 'Email Taken')
				return redirect('register')
			else:
				user = User.objects.create_user(username=username, password=password, email=email, first_name=firstname, last_name=lastname)
				user.save()
				messages.info(request, 'User Created')
		
				for label in filelabel:
					currentid = []
					
					for rule in rules:
						CurrentRule = json.loads(rule.rule)
						if CurrentRule["module"] == label:
							currentid.append(rule.ruleid)
					
					for rid in currentid:
						mapping = FilesData(filelabel=label, userid=user.id, ruleid=rid)
						mapping.save()

				return redirect('register')
		else:
			messages.info(request, 'Password Not Matching')
			return redirect('register')
		return redirect('/')
	else:
		return render(request, 'register.html', {'rules': ruleList })

def uploadrules(request):
	rulesAll = Rules.objects.all()


	if request.method =="POST":
		rulelabel = request.POST['rulelabel']
		pattern_name = request.POST['pattern_name']
		sheetName = request.POST['sheetName']
		header = request.POST['header']
		comma_columns = request.POST['columnName']
		columns = comma_columns.split(",")

		newrule = {}
		newrule["ruleId"]=8
		newrule["rulelabel"]=rulelabel
		newrule["pattern_name"]=pattern_name
		newrule["sheetName"]=sheetName
		newrule["header"]=header
		newrule["columns"]=columns
		print(newrule)
		newrule = json.dumps(newrule)

		entry = Rules(rule=newrule)
		entry.save()

		messages.info(request,'Rule Added')
		return redirect('uploadrules') 
		 
	else:
		rulesAllSerialized = serializers.serialize("json", rulesAll)	
		return render(request,'uploadrules.html',{'rulesAllSerialized':rulesAllSerialized}) 

@csrf_exempt
def update_rules(request):
	if request.method == 'POST':

		
		rules = Rules.objects.filter(ruleid=request.POST.get('ruleid'))
		
		ruleid= request.POST.get('ruleid')
		rulelabel = request.POST.get('rulelabel')
		pattern_name = request.POST.get('updtpname')
		period = request.POST.get('updtpperiod')
		sheetName = request.POST.get('updtpsheet')
		header = request.POST.get('updtpheader')
		comma_columns = request.POST.get('updtpcolumns')
		columns = comma_columns.split(",")
        
		newrule = {}
		newrule["rulelabel"]=rulelabel
		newrule["pattern_name"]=pattern_name
		newrule["pattern_period"]=period
		newrule["pattern_sheetName"]=sheetName
		newrule["header"]=header
		newrule["columns"]=columns
		
		newrule = json.dumps(newrule)
		print("##############")
		print(newrule)
		print(ruleid)
		print("##############")

		rules.rule = newrule
		rules.rule
		rules.save()

		messages.info(request,'Rule Updated')
		return redirect('uploadrules') 