import re
import xlrd 
import os 
import json
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
FILES_FOLDER = os.path.join(BASE_DIR,'code\\Media\\')
SCRIPTS_FOLDER = os.path.join(BASE_DIR,'code\\Scripts\\')

# Read config json file and store in an object
f = open(os.path.join(SCRIPTS_FOLDER, 'metadata.json'))
json_string = f.read()
config = json.loads(json_string)
f.close()

filename = "AP&ISA_202002Feb_Maximo_Resoln_wos_BI_20200310.xlsx" # will flow from the frontend as input
filepath = os.path.join(FILES_FOLDER, filename)

ruleid = 2 # will flow from frontend as input
rule = config[ruleid]

#1st check if the file name is correct with pattern name and patter period matching strings
if re.search(rule["pattern_period"],filename,re.IGNORECASE) and re.search(rule["pattern_name"],filename,re.IGNORECASE):
	pattern_period = (re.search(rule["pattern_period"],filename,re.IGNORECASE)).group(0) #extract the period 
	book = openpyxl.load_workbook(filepath) # if yes, open the workbook 
	
	for sheetname in book.sheetnames:
		if re.search(rule["sheetName"],sheetname,re.IGNORECASE): # check if we have the required sheet. 
			xl_sheet = book[sheetname]
			if xl_sheet.max_column == 0:
				print ("no data")
			else:
				columns=[]
				if rule["header"]:
					header = rule["header"]				
					for cell in xl_sheet[header]: 
						columns.append(cell.value)
					#now check for missing columns 						
					missing_columns =  [item for item in rule["columns"] if item not in columns]

					if missing_columns:
						print ("missing columns")
					else:
						print ("valid file")
					
				else:
					no_of_rows = xl_sheet.max_row
					#searching top 50 rows for headers
					counter = 1
					while counter < no_of_rows:
						for cell in xl_sheet[counter]: 
							columns.append(cell.value)
						if [item for item in rule["columns"] if item not in columns]:
							counter = counter + 1 
							continue
						else:
							print ("found header: ",counter," valid file") # in this case we must return header value also
							break
						
else:
	print("File Name not correct")
 