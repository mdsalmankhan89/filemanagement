import openpyxl
import os
import re 
import psycopg2
import json


 
def validateFileExtension(filepath,validationResult):

    extension  = os.path.splitext(filepath)[1]

    #check if extension is correct
    if re.search(".xlsx",extension,re.IGNORECASE):
        validationResult.update({'extension':True})
    else:
        validationResult.update({'extension':False})

    return validationResult

def validateFileNamePattern(rules,filepath,validationResult,module):
    filename = os.path.basename(filepath)
    validationResult.update({'filename_pattern':["",filename,False]})
    
    for rule in rules:
        ruleJSON = json.loads(rule[1])
        if re.search(ruleJSON["filename_pattern"],filename,re.IGNORECASE) and re.search(ruleJSON["module"],module,re.IGNORECASE):
            validationResult.update({'filename_pattern':[ruleJSON["filename_pattern"],filename,True]})
    
    return validationResult

def validatePeriodPattern(rules,filepath,validationResult,module):
    filename = os.path.basename(filepath)
    validationResult.update({'period_pattern':["",filename,"",False]})

    for rule in rules:
        ruleJSON = json.loads(rule[1])

        if re.search(ruleJSON["filename_pattern"],filename,re.IGNORECASE) and re.search(ruleJSON["module"],module,re.IGNORECASE):

            if re.search(ruleJSON["period_pattern"],filename,re.IGNORECASE):
                validationResult.update({'period_pattern':[ruleJSON["period_pattern"],filename,(re.search(ruleJSON["period_pattern"],filename,re.IGNORECASE)).group(0),True]})
                break
            else:
                validationResult.update({'period_pattern':[ruleJSON["period_pattern"],filename,"",False]})

    return validationResult

def validateSheets(rules,filepath,validationResult,module):
    expectedsheets = {}

    book = openpyxl.load_workbook(filepath)
    presentSheets = [x.lower() for x in book.sheetnames]

    for rule in rules:
        ruleJSON = json.loads(rule[1])

        if (re.search(ruleJSON["module"],module,re.IGNORECASE) and (ruleJSON["period_pattern"]==validationResult["period_pattern"][0]) and re.search(ruleJSON["filename_pattern"],validationResult["filename_pattern"][0],re.IGNORECASE)):
            if ([item for item in [(ruleJSON["sheetname_pattern"]).lower()] if item in presentSheets]):
                expectedsheets.update({ruleJSON["sheetname_pattern"]:True})
            else:
                expectedsheets.update({ruleJSON["sheetname_pattern"]:False})
    if ([item for item in [False] if item in expectedsheets.values()] or (not expectedsheets)):
        validationResult.update({'sheetname_pattern':[expectedsheets,False]})
    else:
        validationResult.update({'sheetname_pattern':[expectedsheets,True]})
    
    return validationResult



def validateColumns(rules,filepath,validationResult,module):
    expectedresult = {}
    book = openpyxl.load_workbook(filepath) 

    for sheetname in book.sheetnames:
        partialmatch_Header=0 
        partialmatch_Counter =0 
        partial_delta = []
        print(sheetname)
        xl_sheet = book[sheetname]
        print(xl_sheet)

        for rule in rules:
            ruleJSON = json.loads(rule[1])
            ruleid = rule[0]
            
			
            if (re.search(ruleJSON["module"],module,re.IGNORECASE) and (ruleJSON["period_pattern"]==validationResult["period_pattern"][0]) and re.search(ruleJSON["filename_pattern"],validationResult["filename_pattern"][0],re.IGNORECASE) and re.search(ruleJSON["sheetname_pattern"],sheetname,re.IGNORECASE)):
                print("Entry")
                if ruleJSON["header"] and ruleJSON["column_offset"]:
                    columns=[]
                    for cell in xl_sheet[ruleJSON["header"]]:
                        columns.append(cell.value)
                    missing_columns =  [item for item in ruleJSON["columns"] if item not in columns]
                    print("yes")
                    print(columns)
                    print(missing_columns)

                    if missing_columns:
                        expectedresult.update({sheetname:[ruleJSON["columns"],missing_columns,ruleJSON["header"],ruleJSON["column_offset"],False,None]})
                    else:
                        expectedresult.update({sheetname:[ruleJSON["columns"],None,ruleJSON["header"],ruleJSON["column_offset"],True,ruleid]})
                        if int(ruleJSON["column_offset"]) > 0:
                            xl_sheet.delete_cols(1, int(ruleJSON["column_offset"]))
                        if int(ruleJSON["header"])>1 :
                            xl_sheet.delete_rows(1,(int(ruleJSON["header"])-1))
                
                if (ruleJSON["header"] == "" or ruleJSON["header"] is None):
                    no_of_rows = xl_sheet.max_row
                    counter = 1
                    while counter < no_of_rows:
                        columns=[]
                        for cell in xl_sheet[counter]:
                            columns.append(cell.value)
                            print(columns)
							
                        if [item for item in [ruleJSON["columns"]] if item in columns]:
                            if counter==1:
                                partialmatch_Counter = len([item for item in [ruleJSON["columns"]] if item in columns])
                                partialmatch_Header = counter
                                partial_delta = [item for item in [ruleJSON["columns"]] if item in columns]
                                partial_col_offset = 0
                                for val in columns:
                                    if val is None:
                                        partial_col_offset += 1
                            elif  partialmatch_Counter > len([item for item in [ruleJSON["columns"]] if item in columns]):
                                partialmatch_Counter = len([item for item in [ruleJSON["columns"]] if item in columns])
                                partialmatch_Header = counter
                                partial_delta = [item for item in [ruleJSON["columns"]] if item in columns]
                                partial_col_offset = 0
                                for val in columns:
                                    if val is None:
                                        partial_col_offset += 1
                                        
                            counter = counter + 1
                            if counter > 50:
                                break
                            continue
                        else:
                            if ruleJSON["column_offset"]:
                                if col_offset > 0:
                                    xl_sheet.delete_cols(1, ruleJSON["column_offset"])
                                if counter > 1 :
                                    xl_sheet.delete_rows(1,(counter-1))
                                expectedresult.update({sheetname:[ruleJSON["columns"],None,counter,ruleJSON["column_offset"],True,ruleid]})
                                
                            else:
                                col_offset = 0
                                for val in columns:
                                    if val is None:
                                        col_offset += 1
                                if col_offset > 0:
                                    xl_sheet.delete_cols(1, col_offset)
                                if counter > 1 :
                                    xl_sheet.delete_rows(1,(counter-1))
                                expectedresult.update({sheetname:[ruleJSON["columns"],None,counter,col_offset,True,ruleid]})
                            break

                    if counter > 50:
                        print (partialmatch_Counter)
                        print (partialmatch_Header)
                        print (partial_delta )
                        print ("header not found")
                        expectedresult.update({sheetname:[ruleJSON["columns"],partial_delta,partialmatch_Header,partial_col_offset,False,None]})

    print(expectedresult.values())
    result =[]
    for resvals in expectedresult.values():
        result.append(resvals[4])
    
    if ([item for item in [False] if item in result] or (not result)):
        validationResult.update({'columns':[expectedresult,False]})
    else:
        validationResult.update({'columns':[expectedresult,True]})

    return validationResult
                

def validate(filepath,module,rules):

    validationResult  = {}

    validationResult = validateFileExtension(filepath,validationResult)

    if validationResult["extension"] == True:
        
        validationResult = validateFileNamePattern(rules,filepath,validationResult,module)
        print(validationResult)

        if validationResult["filename_pattern"][2]== True:

            validationResult = validatePeriodPattern(rules,filepath,validationResult,module)
            print(validationResult)			

            if validationResult["period_pattern"][3] == True:

                validationResult = validateSheets(rules,filepath,validationResult,module)
                print(validationResult)				

                if validationResult["sheetname_pattern"][1] ==True:

                    validationResult=validateColumns(rules,filepath,validationResult,module)
                    print(validationResult)					

                    if validationResult["columns"][1] == True:
                        print ("success")

def start():
	filepath = "C:\GITHUB\TestFileRules\AP&ISA_202003Mar_Maximo_OpenWO_wos_BI_20200310.xlsx"
	module = "Maximo"
	connection = psycopg2.connect(user="postgres",
                                  password="postgres",
                                  host="127.0.0.1",
                                  port="5432",
                                  database="telusko")
	cursor = connection.cursor()
	postgreSQL_select_Query = "select * from public.account_rules"

	cursor.execute(postgreSQL_select_Query)
	print("Selecting rows from mobile table using cursor.fetchall")
	records = cursor.fetchall()

	validate(filepath,module,records)
		
start()